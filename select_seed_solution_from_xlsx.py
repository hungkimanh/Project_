#!/usr/bin/env python3
import argparse
import pathlib
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook


def _norm_instance(s: str) -> str:
    return pathlib.PurePosixPath((s or "").replace("\\", "/")).as_posix().strip()


def _to_int(v: Any) -> Optional[int]:
    if v is None:
        return None
    try:
        return int(float(v))
    except Exception:
        return None


def _to_float(v: Any) -> Optional[float]:
    if v is None:
        return None
    try:
        return float(v)
    except Exception:
        return None


def _pick_sheet(wb, sheet_name: Optional[str]):
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
        return wb[sheet_name]
    return wb[wb.sheetnames[0]]


def _read_rows(ws) -> Tuple[Dict[str, int], List[Tuple[Any, ...]]]:
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        raise ValueError("Excel sheet is empty")
    col = {str(v).strip(): i for i, v in enumerate(header) if v is not None}
    required = {"instance", "A", "L", "best_solution"}
    missing = required - set(col.keys())
    if missing:
        raise ValueError(f"Missing required columns: {sorted(missing)}")
    return col, list(rows)


def _row_best_fitness(row: Tuple[Any, ...], col: Dict[str, int]) -> float:
    if "best_fitness" not in col:
        return float("inf")
    f = _to_float(row[col["best_fitness"]])
    return f if f is not None else float("inf")


def _row_best_multi_fitness(row: Tuple[Any, ...], col: Dict[str, int]) -> float:
    if "best_multi_fitness" not in col:
        return float("inf")
    f = _to_float(row[col["best_multi_fitness"]])
    return f if f is not None else float("inf")


def _match_instance(excel_instance: str, target_instance: str) -> bool:
    a = _norm_instance(excel_instance)
    b = _norm_instance(target_instance)
    if a == b:
        return True
    return pathlib.PurePosixPath(a).name == pathlib.PurePosixPath(b).name


def main():
    ap = argparse.ArgumentParser(
        description=(
            "Select seed solution from xlsx by (instance, A, L): "
            "prefer best_multi_solution, fallback to best_solution."
        )
    )
    ap.add_argument("--xlsx", required=True, help="Path to xlsx file")
    ap.add_argument("--instance", required=True, help="Instance path to match")
    ap.add_argument("--a", required=True, type=int, help="A value")
    ap.add_argument("--l", required=True, type=int, help="L value")
    ap.add_argument("--seed-out", required=True, help="Output seed file path")
    ap.add_argument("--sheet", default=None, help="Optional sheet name (default: first sheet)")
    args = ap.parse_args()

    xlsx_path = pathlib.Path(args.xlsx)
    if not xlsx_path.exists():
        raise FileNotFoundError(f"xlsx not found: {xlsx_path}")

    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = _pick_sheet(wb, args.sheet)
    col, rows = _read_rows(ws)

    # Candidate tuple:
    # (
    #   source_priority,   # 0 = best_multi_solution, 1 = best_solution fallback
    #   score,             # best_multi_fitness or best_fitness (smaller is better)
    #   row
    # )
    candidates: List[Tuple[int, float, Tuple[Any, ...]]] = []
    for row in rows:
        inst = row[col["instance"]] if col["instance"] < len(row) else None
        if not inst or not _match_instance(str(inst), args.instance):
            continue
        a = _to_int(row[col["A"]] if col["A"] < len(row) else None)
        l = _to_int(row[col["L"]] if col["L"] < len(row) else None)
        if a != args.a or l != args.l:
            continue

        multi_sol = None
        if "best_multi_solution" in col and col["best_multi_solution"] < len(row):
            multi_sol = row[col["best_multi_solution"]]
        if multi_sol and str(multi_sol).strip():
            candidates.append((0, _row_best_multi_fitness(row, col), row))
            continue

        sol = row[col["best_solution"]] if col["best_solution"] < len(row) else None
        if sol and str(sol).strip():
            candidates.append((1, _row_best_fitness(row, col), row))

    if not candidates:
        raise RuntimeError(
            "No usable seed solution found "
            f"(best_multi_solution/best_solution) for "
            f"instance={args.instance}, A={args.a}, L={args.l} in {xlsx_path}"
        )

    candidates.sort(key=lambda x: (x[0], x[1]))
    source_priority, score, best_row = candidates[0]
    if source_priority == 0:
        best_sol = str(best_row[col["best_multi_solution"]]).strip()
        source = "best_multi_solution"
    else:
        best_sol = str(best_row[col["best_solution"]]).strip()
        source = "best_solution"

    seed_path = pathlib.Path(args.seed_out)
    seed_path.parent.mkdir(parents=True, exist_ok=True)
    seed_path.write_text("solution = " + best_sol + "\n", encoding="utf-8")

    print(
        "[SEED_PICK] "
        f"instance={args.instance} A={args.a} L={args.l} "
        f"source={source} fitness={score} seed={seed_path}"
    )


if __name__ == "__main__":
    main()
