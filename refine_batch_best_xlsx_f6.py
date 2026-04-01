#!/usr/bin/env python3
import argparse
import ast
import csv
import json
import os
import pathlib
import subprocess
import sys
import tempfile
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook


def _to_int(x: Any) -> Optional[int]:
    try:
        return int(x)
    except Exception:
        try:
            return int(float(x))
        except Exception:
            return None


def read_instance_releases(path: str) -> List[int]:
    lines = pathlib.Path(path).read_text(encoding="utf-8", errors="ignore").splitlines()
    if not lines:
        return []
    first = lines[0].strip().split()
    parameterized = bool(first) and first[0] == "number_truck"
    header_only = (not parameterized) and bool(first) and (not _is_numeric(first[0]))
    if parameterized:
        idx = 8
    elif header_only:
        idx = 1
    else:
        idx = 0
    releases: List[int] = []
    for ln in lines[idx:]:
        tok = ln.strip().split()
        if len(tok) < 4:
            continue
        rel = _to_int(tok[3])
        if rel is None:
            rel = 0
        releases.append(rel)
    return releases


def _is_numeric(s: str) -> bool:
    try:
        float(s)
        return True
    except Exception:
        return False


def normalize_solution_text(solution_text: str, releases: List[int]) -> str:
    st = (solution_text or "").strip()
    if not st:
        return ""
    try:
        data = ast.literal_eval(st)
    except Exception:
        return st
    if not isinstance(data, list) or len(data) != 2:
        return st
    trucks_raw = data[0]
    drone_raw = data[1]
    if not isinstance(trucks_raw, list) or not isinstance(drone_raw, list):
        return st

    n = len(releases)
    owner: Dict[int, int] = {}
    pos_on_truck: Dict[int, int] = {}
    trucks: List[List[List[Any]]] = []
    for tid, route in enumerate(trucks_raw):
        if not isinstance(route, list):
            route = []
        route_norm: List[List[Any]] = []
        for pos, stop in enumerate(route):
            if not isinstance(stop, (list, tuple)) or len(stop) < 2:
                continue
            city = _to_int(stop[0])
            pkgs = stop[1] if isinstance(stop[1], list) else []
            if city is None:
                continue
            pkgs_norm = [int(x) for x in pkgs if _to_int(x) is not None]
            route_norm.append([city, pkgs_norm])
            if city > 0 and city < n:
                owner[city] = tid
                pos_on_truck[city] = pos
        trucks.append(route_norm)

    drone: List[List[List[Any]]] = []
    for trip in drone_raw:
        if not isinstance(trip, list):
            continue
        trip_norm: List[List[Any]] = []
        for ev in trip:
            if not isinstance(ev, (list, tuple)) or len(ev) < 2:
                continue
            rv = _to_int(ev[0])
            pkgs = ev[1] if isinstance(ev[1], list) else []
            if rv is None:
                continue
            pkgs_norm = [int(x) for x in pkgs if _to_int(x) is not None]
            trip_norm.append([rv, pkgs_norm])
        drone.append(trip_norm)

    def eligible(pkg: int) -> bool:
        return 0 < pkg < n and releases[pkg] > 0

    while True:
        changed = False
        cleaned_drone: List[List[List[Any]]] = []
        for trip in drone:
            kept_events: List[List[Any]] = []
            for ev in trip:
                rv = _to_int(ev[0])
                pkgs = ev[1] if isinstance(ev[1], list) else []
                if rv is None or rv <= 0 or rv >= n:
                    changed = True
                    continue
                tid = owner.get(rv, -1)
                if tid < 0:
                    changed = True
                    continue
                rv_pos = pos_on_truck.get(rv, -1)
                if rv_pos < 0:
                    changed = True
                    continue
                kept_pkgs: List[int] = []
                for x in pkgs:
                    px = _to_int(x)
                    if px is None:
                        changed = True
                        continue
                    pkg = int(px)
                    if not eligible(pkg):
                        changed = True
                        continue
                    if owner.get(pkg, -1) != tid:
                        changed = True
                        continue
                    if pos_on_truck.get(pkg, -1) < rv_pos:
                        changed = True
                        continue
                    kept_pkgs.append(pkg)
                if not kept_pkgs:
                    changed = True
                    continue
                kept_events.append([rv, kept_pkgs])
            if kept_events:
                cleaned_drone.append(kept_events)
            else:
                if trip:
                    changed = True
        drone = cleaned_drone

        is_resup = [0] * n
        for trip in drone:
            for ev in trip:
                for pkg in ev[1]:
                    if 0 < pkg < n:
                        is_resup[pkg] = 1

        truck_count = max((tid for tid in owner.values()), default=-1) + 1
        rmax = [0] * max(1, truck_count)
        for c in range(1, n):
            tid = owner.get(c, -1)
            if tid < 0:
                continue
            if not is_resup[c]:
                rmax[tid] = max(rmax[tid], releases[c])

        pruned_drone: List[List[List[Any]]] = []
        removed_by_rmax = False
        for trip in drone:
            kept_events: List[List[Any]] = []
            for ev in trip:
                rv = ev[0]
                tid = owner.get(rv, -1)
                if tid < 0:
                    removed_by_rmax = True
                    continue
                kept_pkgs = []
                for pkg in ev[1]:
                    if pkg <= rmax[tid]:
                        removed_by_rmax = True
                        continue
                    kept_pkgs.append(pkg)
                if kept_pkgs:
                    kept_events.append([rv, kept_pkgs])
                else:
                    removed_by_rmax = True
            if kept_events:
                pruned_drone.append(kept_events)
        drone = pruned_drone

        if not changed and not removed_by_rmax:
            break

    out = [trucks, drone]
    return json.dumps(out, separators=(",", ":"))


def sheet_to_rows(xlsx_path: pathlib.Path, sheet_name: Optional[str]) -> Tuple[List[str], List[Dict[str, Any]]]:
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    header_row = next(it, None)
    if not header_row:
        raise ValueError("Empty sheet")
    header = [str(x) if x is not None else "" for x in header_row]
    rows: List[Dict[str, Any]] = []
    for r in it:
        row = {}
        for i, c in enumerate(header):
            row[c] = r[i] if i < len(r) else None
        rows.append(row)
    return header, rows


def write_csv(path: pathlib.Path, fieldnames: List[str], rows: List[Dict[str, Any]]) -> None:
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for row in rows:
            out = {}
            for c in fieldnames:
                v = row.get(c, "")
                out[c] = "" if v is None else str(v)
            w.writerow(out)


def read_csv(path: pathlib.Path) -> Tuple[List[str], List[Dict[str, str]]]:
    with path.open(newline="", encoding="utf-8") as f:
        r = csv.DictReader(f)
        rows = list(r)
        return (r.fieldnames or []), rows


def write_xlsx(path: pathlib.Path, sheet_name: str, fieldnames: List[str], rows: List[Dict[str, str]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(fieldnames)
    for row in rows:
        ws.append([row.get(c, "") for c in fieldnames])
    wb.save(path)


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--in-xlsx", required=True)
    ap.add_argument("--out-xlsx", required=True)
    ap.add_argument("--sheet", default="yellow_rows")
    ap.add_argument("--solver-bin", default="C_Version/read_data_f6")
    ap.add_argument("--workers", type=int, default=8)
    args = ap.parse_args()

    in_xlsx = pathlib.Path(args.in_xlsx)
    out_xlsx = pathlib.Path(args.out_xlsx)
    fieldnames, rows = sheet_to_rows(in_xlsx, args.sheet)

    required = {"instance", "A", "L", "best_solution", "best_multi_solution"}
    missing = required - set(fieldnames)
    if missing:
        raise ValueError(f"Missing columns: {sorted(missing)}")

    # No Python-side normalization here.
    # Keep Function6.cpp as the single source of truth for normalize/validate/fitness.

    with tempfile.TemporaryDirectory(prefix="refine_xlsx_") as td:
        td_path = pathlib.Path(td)
        tmp_csv = td_path / "tmp.csv"
        write_csv(tmp_csv, fieldnames, rows)

        env = os.environ.copy()
        env["SOLVER_BIN"] = args.solver_bin
        env["PARALLEL_JOBS"] = str(args.workers)
        p = subprocess.run(
            [sys.executable, "recompute_metrics_from_solutions_csv.py", str(tmp_csv)],
            text=True,
            env=env,
        )
        if p.returncode != 0:
            return p.returncode

        out_fields, out_rows = read_csv(tmp_csv)
        write_xlsx(out_xlsx, args.sheet, out_fields, out_rows)
        print(f"Wrote refined xlsx: {out_xlsx}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
