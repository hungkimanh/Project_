import Data
import Function
import Neighborhood
import time
import copy
import random
import Neighborhood11
import Neighborhood10
import Neighborhood_drone
import glob
import os
import openpyxl
import csv
import numpy as np
import math
import sys
global LOOP
global tabu_tenure
global best_sol
global best_fitness
global Tabu_Structure
global current_neighborhood
global LOOP_IMPROVED
global SET_LAST_10
global BEST
global start_time

# Set up chỉ số -------------------------------------------------------------------
ITE = 10
epsilon = (-1) * 0.00001
# 15:   120,    20:    150
# BREAKLOOP = Data.number_of_cities * 8
LOOP_IMPROVED = 0
SET_LAST_10 = [] 
BEST = []
TIME_LIMIT = 3000
data_set = os.getenv('DATA_SET', "U_10_0.5_Num_1.txt")
center = os.getenv('CENTER', "center")
start_time = time.time()
# 
solution_pack_len = 5
# random.seed(7)

def roulette_wheel_selection(population, fitness_scores): 
    total_fitness = sum(fitness_scores)
    probabilities = [score / total_fitness for score in fitness_scores]
    selected_index = np.random.choice(len(population), p=probabilities)
    return population[selected_index]

def Tabu_search(init_solution, tabu_tenure, CC, first_time, Data1, index_consider_elite_set, solution_pack):
    current_fitness, current_truck_time, current_sum_fitness = Function.fitness(init_solution)
    best_sol = init_solution
    best_fitness = current_fitness
    sol_chosen_to_break = init_solution
    fit_of_sol_chosen_to_break = current_fitness
    
    lennn = [0] * 6
    lenght_i = [0] * 6
    i = 0
    
    Result_print = []
    # LOOP = BREAKLOOP * AA
    # print(Data.standard_deviation)T +=
    global current_neighborhood
    global LOOP_IMPROVED
    LOOP_IMPROVED = 0
    global use_optimize_truck_route
    use_optimize_truck_route = False
    
    Data1 = [['act', 'fitness', 'change1', 'change2', 'solution', 'tabu structue', 'tabu structure1']]
    LOOP = min(int(Data.number_of_cities*math.log(Data.number_of_cities)), 100)

    # BREAKLOOP = Data.number_of_cities
    SEGMENT = 5
    END_SEGMENT =  int(Data.number_of_cities/math.log10(Data.number_of_cities))*2
    
    T = 0
    nei_set = [0, 1, 2, 3, 4, 5, 6]
    weight = [1/len(nei_set)]*len(nei_set)
    current_sol = init_solution
    
    while T < SEGMENT:
        end_time = time.time()
        if end_time - start_time > TIME_LIMIT:
            break
        tabu_tenure = tabu_tenure1 = tabu_tenure3 = tabu_tenure2 = random.uniform(2*math.log(Data.number_of_cities), Data.number_of_cities)
        Tabu_Structure = [(tabu_tenure +1) * (-1)] * Data.number_of_cities
        Tabu_Structure1 = [(tabu_tenure +1) * (-1)] * Data.number_of_cities
        Tabu_Structure2 = [(tabu_tenure +1) * (-1)] * Data.number_of_cities
        Tabu_Structure3 = [(tabu_tenure +1) * (-1)] * Data.number_of_cities
        factor = 0.3 #0.3 0.6
        score = [0]*len(nei_set)
        used = [0]*len(nei_set)
        prev_f = best_fitness
        
        LOOP_IMPROVED = 0
        lennn = [0] * 6
        lenght_i = [0] * 6
        i = 0
        while i < END_SEGMENT:
            prev_fitness = current_fitness
            current_neighborhood = []
            choose = roulette_wheel_selection(nei_set, weight)
            if choose == 0:
                current_neighborhood1, solution_pack = Neighborhood.Neighborhood_combine_truck_and_drone_neighborhood_with_tabu_list_with_package(name_of_truck_neiborhood=Neighborhood10.Neighborhood_one_opt_standard, solution=current_sol, number_of_potial_solution=CC, number_of_loop_drone=2, tabu_list=Tabu_Structure, tabu_tenure=tabu_tenure,  index_of_loop=lenght_i[1], best_fitness=best_fitness, kind_of_tabu_structure=1, need_truck_time=False, solution_pack=solution_pack, solution_pack_len=solution_pack_len, use_solution_pack=first_time, index_consider_elite_set=index_consider_elite_set)
                current_neighborhood.append([1, current_neighborhood1])
            elif choose == 1:
                current_neighborhood1, solution_pack = Neighborhood.Neighborhood_combine_truck_and_drone_neighborhood_with_tabu_list_with_package(name_of_truck_neiborhood=Neighborhood10.Neighborhood_one_opt_standard, solution=current_sol, number_of_potial_solution=CC, number_of_loop_drone=2, tabu_list=Tabu_Structure, tabu_tenure=tabu_tenure,  index_of_loop=lenght_i[1], best_fitness=best_fitness, kind_of_tabu_structure=1, need_truck_time=False, solution_pack=solution_pack, solution_pack_len=solution_pack_len, use_solution_pack=first_time, index_consider_elite_set=index_consider_elite_set)
                current_neighborhood.append([2, current_neighborhood1])
            elif choose == 2:
                current_neighborhood5, solution_pack = Neighborhood.Neighborhood_combine_truck_and_drone_neighborhood_with_tabu_list_with_package(name_of_truck_neiborhood=Neighborhood11.Neighborhood_two_opt_tue, solution=current_sol, number_of_potial_solution=CC, number_of_loop_drone=2, tabu_list=Tabu_Structure3, tabu_tenure=tabu_tenure3,  index_of_loop=lenght_i[5], best_fitness=best_fitness, kind_of_tabu_structure=5, need_truck_time=False, solution_pack=solution_pack, solution_pack_len=solution_pack_len, use_solution_pack=first_time, index_consider_elite_set=index_consider_elite_set)
                current_neighborhood.append([5, current_neighborhood5])
            elif choose == 3: 
                current_neighborhood4, solution_pack = Neighborhood.Neighborhood_combine_truck_and_drone_neighborhood_with_tabu_list_with_package(name_of_truck_neiborhood=Neighborhood11.Neighborhood_move_2_1, solution=current_sol, number_of_potial_solution=CC, number_of_loop_drone=2, tabu_list=Tabu_Structure2, tabu_tenure=tabu_tenure2,  index_of_loop=lenght_i[4], best_fitness=best_fitness, kind_of_tabu_structure=4, need_truck_time=False, solution_pack=solution_pack, solution_pack_len=solution_pack_len, use_solution_pack=first_time, index_consider_elite_set=index_consider_elite_set)
                current_neighborhood.append([4, current_neighborhood4])
            elif choose == 4: 
                current_neighborhood3, solution_pack = Neighborhood.Neighborhood_combine_truck_and_drone_neighborhood_with_tabu_list_with_package(name_of_truck_neiborhood=Neighborhood11.Neighborhood_move_1_1_standard, solution=current_sol, number_of_potial_solution=CC, number_of_loop_drone=2, tabu_list=Tabu_Structure1, tabu_tenure=tabu_tenure1,  index_of_loop=lenght_i[3], best_fitness=best_fitness, kind_of_tabu_structure=3, need_truck_time=False, solution_pack=solution_pack, solution_pack_len=solution_pack_len, use_solution_pack=first_time, index_consider_elite_set=index_consider_elite_set)
                current_neighborhood.append([3, current_neighborhood3])
            elif choose == 5: 
                current_neighborhood3, solution_pack = Neighborhood.Neighborhood_combine_truck_and_drone_neighborhood_with_package(name_of_truck_neiborhood=Neighborhood.Neighborhood_split_two_truck_term, solution=current_sol, number_of_potiential_solution=CC, number_of_loop_drone=2, whether_use_truck_time=False, solution_pack=solution_pack, solution_pack_len=solution_pack_len, use_solution_pack=first_time, index_consider_elite_set=index_consider_elite_set)
                current_neighborhood.append([6, current_neighborhood3])
            else:
                current_neighborhood3, solution_pack = Neighborhood.Neighborhood_combine_truck_and_drone_neighborhood_with_package(name_of_truck_neiborhood=Neighborhood.Neighborhood_stack_two_truck_term, solution=current_sol, number_of_potiential_solution=CC, number_of_loop_drone=2, whether_use_truck_time=False, solution_pack=solution_pack, solution_pack_len=solution_pack_len, use_solution_pack=first_time, index_consider_elite_set=index_consider_elite_set)
                current_neighborhood.append([7, current_neighborhood3])

            flag = False
            index = [0] * len(current_neighborhood)
            min_nei = [100000] * len(current_neighborhood)
            min_sum = [1000000000] * len(current_neighborhood)
            # print(current_neighborhood)
            for j in range(len(current_neighborhood)):
                if current_neighborhood[j][0] in [1, 2]:
                    for k in range(len(current_neighborhood[j][1])):
                        cfnode = current_neighborhood[j][1][k][1][0]
                        if cfnode - best_fitness < epsilon:
                            min_nei[j] = cfnode
                            index[j] = k
                            best_fitness = cfnode
                            best_sol = current_neighborhood[j][1][k][0]
                            LOOP_IMPROVED = i
                            flag = True

                        elif cfnode - min_nei[j] < epsilon and Tabu_Structure[current_neighborhood[j][1][k][2]] + tabu_tenure <= lenght_i[1]:
                            min_nei[j] = cfnode
                            index[j] = k
                            min_sum[j] = current_neighborhood[j][1][k][1][2]

                        elif min_nei[j] - epsilon > cfnode and Tabu_Structure[current_neighborhood[j][1][k][2]] + tabu_tenure <= lenght_i[1]:
                            if min_sum[j] > current_neighborhood[j][1][k][1][2]:
                                min_nei[j] = cfnode
                                index[j] = k
                                min_sum[j] = current_neighborhood[j][1][k][1][2]
                elif current_neighborhood[j][0] == 3:
                    for k in range(len(current_neighborhood[j][1])):    
                        cfnode = current_neighborhood[j][1][k][1][0]
                        if cfnode - best_fitness < epsilon:
                            min_nei[j] = cfnode
                            index[j] = k
                            best_fitness = cfnode
                            best_sol = current_neighborhood[j][1][k][0]
                            LOOP_IMPROVED = i
                            flag = True

                        elif cfnode - min_nei[j] < epsilon and Tabu_Structure1[current_neighborhood[j][1][k][2][0]] + tabu_tenure1 <= lenght_i[3] or Tabu_Structure1[current_neighborhood[j][1][k][2][1]] + tabu_tenure1 <= lenght_i[3]:
                            min_nei[j] = cfnode
                            index[j] = k
                            min_sum[j] = current_neighborhood[j][1][k][1][2]

                        elif cfnode < min_nei[j] - epsilon and Tabu_Structure1[current_neighborhood[j][1][k][2][0]] + tabu_tenure1 <= lenght_i[3] or Tabu_Structure1[current_neighborhood[j][1][k][2][1]] + tabu_tenure1 <= lenght_i[3]:
                            if min_sum[j] > current_neighborhood[j][1][k][1][2]:
                                min_nei[j] = cfnode
                                index[j] = k
                                min_sum[j] = current_neighborhood[j][1][k][1][2]
                elif current_neighborhood[j][0] == 4:
                    for k in range(len(current_neighborhood[j][1])):
                        cfnode = current_neighborhood[j][1][k][1][0]
                        if cfnode - best_fitness < epsilon:
                            min_nei[j] = cfnode
                            index[j] = k
                            best_fitness = cfnode
                            best_sol = current_neighborhood[j][1][k][0]
                            LOOP_IMPROVED = i
                            flag = True

                        elif cfnode - min_nei[j] < epsilon and Tabu_Structure2[current_neighborhood[j][1][k][2][0]] + tabu_tenure2 <= lenght_i[4] or Tabu_Structure2[current_neighborhood[j][1][k][2][1]] + tabu_tenure2 <= lenght_i[4] or Tabu_Structure2[current_neighborhood[j][1][k][2][2]] + tabu_tenure2 <= lenght_i[4]:
                            min_nei[j] = cfnode
                            index[j] = k
                            min_sum[j] = current_neighborhood[j][1][k][1][2]
                            
                        elif cfnode < min_nei[j] - epsilon and Tabu_Structure2[current_neighborhood[j][1][k][2][0]] + tabu_tenure2 <= lenght_i[4] or Tabu_Structure2[current_neighborhood[j][1][k][2][1]] + tabu_tenure2 <= lenght_i[4] or Tabu_Structure2[current_neighborhood[j][1][k][2][2]] + tabu_tenure2 <= lenght_i[4]:
                            if min_sum[j] > current_neighborhood[j][1][k][1][2]:
                                min_nei[j] = cfnode
                                index[j] = k
                                min_sum[j] = current_neighborhood[j][1][k][1][2]
                elif current_neighborhood[j][0] == 5:
                    for k in range(len(current_neighborhood[j][1])):    
                        cfnode = current_neighborhood[j][1][k][1][0]
                        if cfnode - best_fitness < epsilon:
                            min_nei[j] = cfnode
                            index[j] = k
                            best_fitness = cfnode
                            best_sol = current_neighborhood[j][1][k][0]
                            LOOP_IMPROVED = i
                            flag = True

                        elif cfnode - min_nei[j] < epsilon and Tabu_Structure3[current_neighborhood[j][1][k][2][0]] + tabu_tenure3 <= lenght_i[5] or Tabu_Structure3[current_neighborhood[j][1][k][2][1]] + tabu_tenure3 <= lenght_i[5]:
                            min_nei[j] = cfnode
                            index[j] = k
                            min_sum[j] = current_neighborhood[j][1][k][1][2]

                        elif cfnode < min_nei[j] - epsilon and Tabu_Structure3[current_neighborhood[j][1][k][2][0]] + tabu_tenure3 <= lenght_i[5] or Tabu_Structure3[current_neighborhood[j][1][k][2][1]] + tabu_tenure3 <= lenght_i[5]:
                            if min_sum[j] > current_neighborhood[j][1][k][1][2]:
                                min_nei[j] = cfnode
                                index[j] = k
                                min_sum[j] = current_neighborhood[j][1][k][1][2]
                else:
                    for k in range(len(current_neighborhood[j][1])):
                        cfnode = current_neighborhood[j][1][k][1][0]
                        if cfnode - best_fitness < epsilon:
                            min_nei[j] = cfnode
                            index[j] = k
                            best_fitness = cfnode
                            best_sol = current_neighborhood[j][1][k][0]
                            LOOP_IMPROVED = i
                            flag = True
                            
                        elif cfnode - min_nei[j] < epsilon:
                            min_nei[j] = cfnode
                            index[j] = k
                            min_sum[j] = current_neighborhood[j][1][k][1][2]
                            
                        elif cfnode < min_nei[j] - epsilon:
                            if min_sum[j] > current_neighborhood[j][1][k][1][2]:
                                min_nei[j] = cfnode
                                index[j] = k
                                min_sum[j] = current_neighborhood[j][1][k][1][2]
            index_best_nei = 0
            best_fit_in_cur_loop = min_nei[0]
            
            # for j in range(len(min_nei)):
            #     print(min_nei[j])
            #     print(current_neighborhood[j][1][index[j]][0])
            #     print("-------")
            
            for j in range(1, len(min_nei)):
                if min_nei[j] < best_fit_in_cur_loop:
                    index_best_nei = j
                    best_fit_in_cur_loop = min_nei[j]
            
            if current_neighborhood[index_best_nei][0] in [1, 2]:
                lenght_i[1] += 1
            
            if current_neighborhood[index_best_nei][0] == 3:
                lenght_i[3] += 1
                
            if current_neighborhood[index_best_nei][0] == 4:
                lenght_i[4] += 1
                
            if current_neighborhood[index_best_nei][0] == 5:
                lenght_i[5] += 1
                
            # print(current_neighborhood[index_best_nei][0])
            # print(len(current_neighborhood[index_best_nei][1]))
            # print(current_neighborhood[index_best_nei][1])
            # print(lenght_i[1], " then ", Tabu_Structure)
            # print(lenght_i[3], " then ", Tabu_Structure1)
            # print(lenght_i[4], " then ", Tabu_Structure2)
            # print(lenght_i[5], " then ", Tabu_Structure3)

            if len(current_neighborhood[index_best_nei][1]) == 0:
                # print("hahhaa")
                continue
                
            # print(index[index_best_nei])
            current_sol = current_neighborhood[index_best_nei][1][index[index_best_nei]][0]
            current_fitness = current_neighborhood[index_best_nei][1][index[index_best_nei]][1][0]
            current_truck_time = current_neighborhood[index_best_nei][1][index[index_best_nei]][1][1]
            current_sum_fitness = current_neighborhood[index_best_nei][1][index[index_best_nei]][1][2]
            
            # print("----------",i,"--------------------------")
            # print(current_sol)
            # print(current_fitness)
            # print(current_neighborhood[index_best_nei][0])
            
            # print(current_fitness, current_sol)
            Data1.append(current_fitness)
            Data1.append(current_sol)
            # SET_LAST_10.append([current_sol, [current_fitness, current_truck_time]])
            # if len(SET_LAST_10) > 10:
            #     SET_LAST_10.pop(0)
            
            if current_neighborhood[index_best_nei][0] in [1, 2]:
                Tabu_Structure[current_neighborhood[index_best_nei][1][index[index_best_nei]][2]] = lenght_i[1] -1
                lennn[current_neighborhood[index_best_nei][0]] += 1
            
            if current_neighborhood[index_best_nei][0] == 3:
                Tabu_Structure1[current_neighborhood[index_best_nei][1][index[index_best_nei]][2][0]] = lenght_i[3] - 1 
                Tabu_Structure1[current_neighborhood[index_best_nei][1][index[index_best_nei]][2][1]] = lenght_i[3] - 1
                lennn[current_neighborhood[index_best_nei][0]] += 1
                
            if current_neighborhood[index_best_nei][0] == 4:
                Tabu_Structure2[current_neighborhood[index_best_nei][1][index[index_best_nei]][2][0]] = lenght_i[4] - 1
                Tabu_Structure2[current_neighborhood[index_best_nei][1][index[index_best_nei]][2][1]] = lenght_i[4] - 1
                Tabu_Structure2[current_neighborhood[index_best_nei][1][index[index_best_nei]][2][2]] = lenght_i[4] - 1
                lennn[current_neighborhood[index_best_nei][0]] += 1
                
            if current_neighborhood[index_best_nei][0] == 5:
                Tabu_Structure3[current_neighborhood[index_best_nei][1][index[index_best_nei]][2][0]] = lenght_i[5] - 1
                Tabu_Structure3[current_neighborhood[index_best_nei][1][index[index_best_nei]][2][1]] = lenght_i[5] - 1
                lennn[current_neighborhood[index_best_nei][0]] += 1
                
            if fit_of_sol_chosen_to_break > current_fitness:
                sol_chosen_to_break = current_sol
                fit_of_sol_chosen_to_break = current_fitness
                LOOP_IMPROVED = i
                
            

            if current_neighborhood[index_best_nei][0] in [1, 2]:
                temp = [current_neighborhood[index_best_nei][0], current_fitness, current_neighborhood[index_best_nei][1][index[index_best_nei]][2], -1, current_sol, Tabu_Structure, Tabu_Structure1]
            elif current_neighborhood[index_best_nei][0] in [3]:
                temp = [current_neighborhood[index_best_nei][0], current_fitness, current_neighborhood[index_best_nei][1][index[index_best_nei]][2][0], current_neighborhood[index_best_nei][1][index[index_best_nei]][2][1], current_sol, Tabu_Structure, Tabu_Structure1]
            else:
                temp = [current_neighborhood[index_best_nei][0], current_fitness, -1, -1, current_sol]
            Data1.append(temp)

            used[choose] += 1
            if flag == True:
                score[choose] += 0.5
            elif current_fitness - prev_fitness < epsilon:
                score[choose] += 0.3
            else:
                score[choose] += 0.1

            for j in range(len(nei_set)):
                if used[j] == 0:
                    continue
                else:
                    weight[j] = (1 - factor)*weight[j] + factor*score[j]/used[j]
            if flag == True:
                i = 0
            else:
                i += 1
        # print("-------",T,"--------")
        # print(best_fitness)
        # print(T, best_sol, "\n", best_fitness)
        # print(used, score, sum(used))

        if best_fitness - prev_f < epsilon:
            T = 0
        else: 
            T += 1
        
    return best_sol, best_fitness, Result_print, solution_pack, Data1
    
def Tabu_search_for_CVRP(CC):
    Data1 = []
    list_init = []
    start_time = time.time()
    solution_pack = []
    current_sol5 = Function.initial_solution7()
    list_init.append(current_sol5)

    
    
    list_fitness_init = []
    fitness5 = Function.fitness(current_sol5)

    list_fitness_init.append(fitness5)

    
    current_fitness = list_fitness_init[0][0]
    current_sol = list_init[0]
    
    for i in range(1, len(list_fitness_init)):
        if current_fitness > list_fitness_init[i][0]:
            current_sol = list_init[i]
            current_fitness = list_fitness_init[i][0]

    # Initial solution thay ở đây ------------->
    # current_sol = check     # Để dòng này làm comment để tìm initial solution theo tham lam
    # <------------- Initial solution thay ở đây 
    
    
    # print(best_sol) 
    # print(best_fitness)
    # print(Function.Check_if_feasible(best_sol))
    solution_pack_len = 5
    best_sol, best_fitness, result_print, solution_pack, Data1 = Tabu_search(init_solution=current_sol, tabu_tenure=Data.number_of_cities-1, CC=CC, first_time=True, Data1=Data1, index_consider_elite_set=0, solution_pack=solution_pack)
    for pi in range(solution_pack_len):
        # print("+++++++++++++++++++++++++",len(solution_pack),"+++++++++++++++++++++++++",)
        # for iiii in range(len(solution_pack)):
        #     print(solution_pack[iiii][0])
        #     print(solution_pack[iiii][1][0])
        #     print("$$$$$$$$$$$$$$")
        if pi < len(solution_pack):
            print("ehehe")
            print(solution_pack[pi][0])
            current_neighborhood5 = Neighborhood.swap_two_array(solution_pack[pi][0])
            best_sol_in_brnei = current_neighborhood5[0][0]
            best_fitness_in_brnei = current_neighborhood5[0][1][0]
            for i in range(1, len(current_neighborhood5)):
                cfnode = current_neighborhood5[i][1][0]
                if cfnode - best_fitness_in_brnei < epsilon:
                    best_sol_in_brnei = current_neighborhood5[i][0]
                    best_fitness_in_brnei = cfnode
            temp = ["break", "break", "break", "break", "break", "break", "break"]
            Data1.append(temp)
            print(best_sol_in_brnei)
            best_sol1, best_fitness1, result_print1, solution_pack, Data1 = Tabu_search(init_solution=best_sol_in_brnei, tabu_tenure=Data.number_of_cities-1, CC=CC, first_time=False, Data1=Data1, index_consider_elite_set=pi+1, solution_pack=solution_pack)
            # print("-----------------", pi, "------------------------")
            # print(best_sol1)
            # print(best_fitness1)
            if best_fitness1 - best_fitness < epsilon:
                best_sol = best_sol1
                best_fitness = best_fitness1
        

    return best_fitness, best_sol

# Thư mục chứa các file .txt
# folder_path = "test_data\\data_demand_random\\10\\"
# folder_path = "test_data\\Smith\\TSPrd(time)\\Solomon\\50\\0_5TSP_50"
# folder_path = "test_data//Smith//TSPrd(time)//Solomon//10//"
# folder_path = "test_data//Smith//TSPrd(time)//Solomon//10//"
folder_path = "test_data//MTVRPRDDR//Instances//"
# Danh sách tất cả các file .txt trong thư mục
txt_files = glob.glob(os.path.join(folder_path, data_set))
# txt_files = glob.glob(os.path.join(folder_path, "*.dat"))
# Tạo một tệp Excel mới
print(txt_files)
workbook = openpyxl.Workbook()
sheet = workbook.active

# Dòng và cột bắt đầu ghi kết quả
row = 1
column = 1

# Ghi tên file .txt vào cột đầu tiên
for txt_file in txt_files:
    sheet.cell(row=row, column=column, value=os.path.basename(txt_file))
    row += 1
# Đặt lại dòng và cột cho việc ghi kết quả
row = 1
for txt_file in txt_files:
    column = 2
    with open(txt_file, 'r') as file:
        # Đọc nội dung từ file .txt và xử lý nó
        # print(txt_file)
        log = os.path.basename(txt_file) + str(Data.number_of_cities) +'.log'
        # log_folder = 'Result/log_result'
        # log_file_path = os.path.join(log_folder, log)
        # log_file = open(log_file_path, 'w')
        # sys.stdout = log_file
        Data.read_data_2024(txt_file, center)
        result = []
        run_time = []
        avg = 0
        avg_run_time = 0
        best_csv_fitness = 1000000
        for i in range(ITE):
            BEST = []
            # print("------------------------",i,"------------------------")
            start_time_1 = time.time()
            best_fitness, best_sol = Tabu_search_for_CVRP(1)
            print("---------- RESULT ----------")
            print(best_sol)
            print(best_fitness)
            avg += best_fitness/ITE
            result.append(best_fitness)
            print(Function.Check_if_feasible(best_sol))
            end_time_1 = time.time()
            run = end_time_1 - start_time_1
            run_time.append(run)
            avg_run_time += run/ITE
            sheet.cell(row=row, column=column, value=best_fitness)

            column += 1
            if best_csv_fitness > best_fitness:
                best_csv_sol = best_sol
                best_csv_fitness = best_fitness
            if i == ITE - 1:
                sheet.cell(row=row, column=column, value=avg_run_time)
                sheet.cell(row=row, column=column+1, value=str(best_csv_sol))
            workbook.save(f"Random_{data_set}_{center}_CL1.xlsx")
        # Tăng dòng cho lần chạy tiếp theo
        row += 1
        
workbook.close()